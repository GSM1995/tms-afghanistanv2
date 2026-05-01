from django.contrib import admin
from django.db.models import Sum, Count
from django.utils.safestring import mark_safe
from django.urls import reverse
from .models import Customer, Driver, Vehicle, Order, ShippingDocument

# برای جلوگیری از خطای تعریف نشدن Invoice در status_badge
from finance.models import Invoice

# ============================================================
# پنل مدیریت مشتریان (Customer)
# ============================================================

@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    # ستون‌های نمایش داده شده در لیست
    list_display = ['avatar_display', 'name', 'phone', 'mobile', 'display_credit', 
                    'order_count', 'total_purchased', 'status_badge', 'action_buttons']
    
    # فیلترهای کناری (فقط فیلدهای واقعی مدل)
    list_filter = ['created_at']  # حذف status_badge چون فیلد واقعی نیست
    
    # فیلدهای جستجو
    search_fields = ['name', 'phone', 'mobile', 'national_id']
    
    # فیلدهای فقط خواندنی
    readonly_fields = ['created_at', 'avatar_preview']
    
    # مرتب‌سازی پیش‌فرض
    ordering = ['-created_at']
    
    # تعداد آیتم در هر صفحه
    list_per_page = 25
    
    # گروه‌بندی فیلدها در فرم افزودن/ویرایش
    fieldsets = (
        ('📋 اطلاعات شخصی', {
            'fields': ('name', 'phone', 'mobile', 'national_id', 'address')
        }),
        ('💰 اطلاعات مالی', {
            'fields': ('credit_limit',),
        }),
        ('👤 اطلاعات کاربری', {
            'fields': ('user',),
            'classes': ('collapse',)
        }),
        ('📅 اطلاعات سیستمی', {
            'fields': ('created_at', 'avatar_preview'),
            'classes': ('collapse',)
        }),
    )
    
    def avatar_display(self, obj):
        """نمایش آواتار/آیکون برای هر مشتری"""
        first_char = obj.name[0] if obj.name else '?'
        colors = ['#3498db', '#27ae60', '#e74c3c', '#f39c12', '#9b59b6', '#1abc9c']
        color_index = hash(obj.name) % len(colors) if obj.name else 0
        color = colors[color_index]
        return mark_safe(f'''
            <div style="
                width: 40px;
                height: 40px;
                background: {color};
                color: white;
                border-radius: 50%;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: bold;
                font-size: 18px;
            ">{first_char}</div>
        ''')
    avatar_display.short_description = ''
    
    def display_credit(self, obj):
        """نمایش اعتبار با رنگ‌بندی و نوار پیشرفت"""
        credit_value = int(obj.credit_limit) if obj.credit_limit else 0
        formatted_credit = f"{credit_value:,}"
        
        if credit_value >= 10000000:
            color = '#27ae60'
            icon = '🟢'
        elif credit_value >= 5000000:
            color = '#f39c12'
            icon = '🟡'
        else:
            color = '#e74c3c'
            icon = '🔴'
        
        percent = min(100, (credit_value / 50000000) * 100)
        
        return mark_safe(f'''
            <div style="min-width: 150px;">
                <div style="display: flex; align-items: center; gap: 5px;">
                    <span style="font-size: 14px;">{icon}</span>
                    <span style="color: {color}; font-weight: bold;">{formatted_credit} ؋</span>
                </div>
                <div style="background: #e9ecef; border-radius: 10px; height: 6px; margin-top: 5px; overflow: hidden;">
                    <div style="background: {color}; width: {percent}%; height: 100%; border-radius: 10px;"></div>
                </div>
            </div>
        ''')
    display_credit.short_description = 'سقف اعتبار'
    
    def order_count(self, obj):
        """تعداد سفارشات مشتری با لینک به سفارشات"""
        count = Order.objects.filter(customer=obj).count()
        url = reverse('admin:orders_order_changelist') + f'?customer__id__exact={obj.id}'
        
        if count == 0:
            return mark_safe(f'<span style="color: #95a5a6;">{count} سفارش</span>')
        return mark_safe(f'<a href="{url}" style="background: #3498db; color: white; padding: 3px 10px; border-radius: 20px; text-decoration: none; font-size: 11px;">📦 {count} سفارش</a>')
    order_count.short_description = 'تعداد سفارشات'
    
    def total_purchased(self, obj):
        """مجموع مبلغ خریداری شده توسط مشتری"""
        total = Order.objects.filter(customer=obj, status='delivered').aggregate(total=Sum('price'))['total'] or 0
        if total == 0:
            return mark_safe('<span style="color: #95a5a6;">-</span>')
        return mark_safe(f'<span style="color: #27ae60; font-weight: bold;">{int(total):,} ؋</span>')
    total_purchased.short_description = 'کل خرید (؋)'
    
    def status_badge(self, obj):
        """نشانگر وضعیت مشتری"""
        order_count = Order.objects.filter(customer=obj).count()
        has_unpaid = Invoice.objects.filter(customer=obj, status='sent').exists()
        
        if has_unpaid:
            return mark_safe('<span style="background: #e74c3c; color: white; padding: 4px 12px; border-radius: 20px; font-size: 11px;">⚠️ بدهکار</span>')
        elif order_count > 0:
            return mark_safe('<span style="background: #27ae60; color: white; padding: 4px 12px; border-radius: 20px; font-size: 11px;">🟢 فعال</span>')
        else:
            return mark_safe('<span style="background: #95a5a6; color: white; padding: 4px 12px; border-radius: 20px; font-size: 11px;">⚪ جدید</span>')
    status_badge.short_description = 'وضعیت'
    
    def action_buttons(self, obj):
        """دکمه‌های عملیات سریع"""
        return mark_safe(f'''
            <div style="display: flex; gap: 5px;">
                <a href="{reverse('admin:orders_customer_change', args=[obj.id])}" 
                   style="background: #3498db; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
                   ✏️
                </a>
                <a href="/customer/login/" 
                   style="background: #27ae60; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
                   👤
                </a>
            </div>
        ''')
    action_buttons.short_description = 'عملیات'
    
    def avatar_preview(self, obj):
        """پیش‌نمایش آواتار در فرم ویرایش"""
        first_char = obj.name[0] if obj.name else '?'
        return mark_safe(f'''
            <div style="
                width: 80px;
                height: 80px;
                background: linear-gradient(135deg, #27ae60, #1a472a);
                color: white;
                border-radius: 50%;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: bold;
                font-size: 36px;
            ">{first_char}</div>
        ''')
    avatar_preview.short_description = 'تصویر نمایه'
    
    # اکشن‌های گروهی
    actions = ['increase_credit', 'export_selected']
    
    def increase_credit(self, request, queryset):
        for customer in queryset:
            customer.credit_limit += 1000000
            customer.save()
        self.message_user(request, f'{queryset.count()} مشتری با موفقیت ۱,۰۰۰,۰۰۰ ؋ افزایش اعتبار یافتند.')
    increase_credit.short_description = 'افزایش ۱ میلیونی اعتبار'
    
    def export_selected(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, Alignment, PatternFill
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customers_export.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "مشتریان"
        
        headers = ['نام', 'تلفن', 'موبایل', 'آدرس', 'اعتبار (؋)', 'تعداد سفارشات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row, customer in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=customer.name)
            ws.cell(row=row, column=2, value=customer.phone)
            ws.cell(row=row, column=3, value=customer.mobile)
            ws.cell(row=row, column=4, value=customer.address)
            ws.cell(row=row, column=5, value=float(customer.credit_limit))
            ws.cell(row=row, column=6, value=Order.objects.filter(customer=customer).count())
        
        wb.save(response)
        return response
    export_selected.short_description = 'خروجی Excel از انتخاب‌ها'


# ============================================================
# پنل مدیریت رانندگان (Driver)
# ============================================================

@admin.register(Driver)
class DriverAdmin(admin.ModelAdmin):
    list_display = ['avatar', 'full_name', 'phone', 'license_number', 'trip_count', 'status_badge', 'action_buttons']
    list_filter = ['is_active']
    search_fields = ['first_name', 'last_name', 'national_code', 'phone']
    
    def avatar(self, obj):
        first_char = obj.first_name[0] if obj.first_name else '?'
        return mark_safe(f'''
            <div style="
                width: 35px;
                height: 35px;
                background: #9b59b6;
                color: white;
                border-radius: 50%;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 14px;
                font-weight: bold;
            ">{first_char}</div>
        ''')
    avatar.short_description = ''
    
    def full_name(self, obj):
        return f"{obj.first_name} {obj.last_name}"
    full_name.short_description = 'نام کامل'
    
    def trip_count(self, obj):
        count = Order.objects.filter(driver=obj).count()
        return mark_safe(f'<span style="background: #3498db; color: white; padding: 2px 8px; border-radius: 20px; font-size: 11px;">🚛 {count} سفر</span>')
    trip_count.short_description = 'تعداد سفرها'
    
    def status_badge(self, obj):
        if obj.is_active:
            return mark_safe('<span style="background: #27ae60; color: white; padding: 3px 10px; border-radius: 20px; font-size: 11px;">✓ فعال</span>')
        return mark_safe('<span style="background: #e74c3c; color: white; padding: 3px 10px; border-radius: 20px; font-size: 11px;">✗ غیرفعال</span>')
    status_badge.short_description = 'وضعیت'
    
    def action_buttons(self, obj):
        return mark_safe(f'''
            <a href="{reverse('admin:orders_driver_change', args=[obj.id])}" 
               style="background: #3498db; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
               ✏️
            </a>
        ''')
    action_buttons.short_description = 'عملیات'


# ============================================================
# پنل مدیریت خودروها (Vehicle)
# ============================================================

@admin.register(Vehicle)
class VehicleAdmin(admin.ModelAdmin):
    list_display = ['plate_number', 'model', 'capacity', 'current_driver', 'availability_badge', 'action_buttons']
    list_filter = ['is_available', 'model']
    search_fields = ['plate_number', 'model']
    
    def availability_badge(self, obj):
        if obj.is_available:
            return mark_safe('<span style="background: #27ae60; color: white; padding: 3px 10px; border-radius: 20px; font-size: 11px;">🚚 در دسترس</span>')
        return mark_safe('<span style="background: #e74c3c; color: white; padding: 3px 10px; border-radius: 20px; font-size: 11px;">⛔ در مسیر</span>')
    availability_badge.short_description = 'وضعیت'
    
    def action_buttons(self, obj):
        return mark_safe(f'''
            <a href="{reverse('admin:orders_vehicle_change', args=[obj.id])}" 
               style="background: #3498db; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
               ✏️
            </a>
        ''')
    action_buttons.short_description = 'عملیات'


# ============================================================
# پنل مدیریت سفارشات (Order)
# ============================================================

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['order_number', 'customer', 'origin', 'destination', 'price', 'status_badge', 'pickup_date', 'action_buttons']
    list_filter = ['status', 'pickup_date']
    search_fields = ['order_number', 'customer__name', 'origin', 'destination']
    date_hierarchy = 'created_at'
    list_per_page = 20
    
    fieldsets = (
        ('اطلاعات اصلی', {
            'fields': ('order_number', 'customer')
        }),
        ('مسیر حمل', {
            'fields': ('origin', 'destination')
        }),
        ('اطلاعات بار', {
            'fields': ('cargo_type', 'weight', 'price', 'advance_payment')
        }),
        ('وضعیت و زمان', {
            'fields': ('status', 'pickup_date', 'driver', 'vehicle')
        }),
    )
    
    def status_badge(self, obj):
        status_colors = {
            'delivered': ('#27ae60', '✅ تحویل شده'),
            'in_transit': ('#3498db', '🚚 در مسیر'),
            'pending': ('#f39c12', '⏳ در انتظار'),
            'assigned': ('#9b59b6', '📋 تخصیص داده شده'),
            'loading': ('#e67e22', '📦 در حال بارگیری'),
            'cancelled': ('#e74c3c', '❌ لغو شده'),
        }
        color, text = status_colors.get(obj.status, ('#95a5a6', obj.status))
        return mark_safe(f'<span style="background: {color}; color: white; padding: 4px 12px; border-radius: 20px; font-size: 11px;">{text}</span>')
    status_badge.short_description = 'وضعیت'
    
    def action_buttons(self, obj):
        return mark_safe(f'''
            <a href="{reverse('admin:orders_order_change', args=[obj.id])}" 
               style="background: #3498db; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
               ✏️
            </a>
        ''')
    action_buttons.short_description = 'عملیات'


# ============================================================
# پنل مدیریت اسناد (ShippingDocument)
# ============================================================

@admin.register(ShippingDocument)
class ShippingDocumentAdmin(admin.ModelAdmin):
    list_display = ['title', 'order', 'document_type_badge', 'uploaded_at', 'action_buttons']
    list_filter = ['document_type']
    search_fields = ['title', 'order__order_number']
    
    def document_type_badge(self, obj):
        type_colors = {
            'waybill': ('#3498db', '📄 بارنامه'),
            'insurance': ('#27ae60', '🛡️ بیمه نامه'),
            'customs': ('#f39c12', '🏛️ گمرکی'),
            'invoice': ('#9b59b6', '🧾 فاکتور'),
            'other': ('#95a5a6', '📎 سایر'),
        }
        color, text = type_colors.get(obj.document_type, ('#95a5a6', obj.document_type))
        return mark_safe(f'<span style="background: {color}; color: white; padding: 3px 10px; border-radius: 20px; font-size: 11px;">{text}</span>')
    document_type_badge.short_description = 'نوع سند'
    
    def action_buttons(self, obj):
        return mark_safe(f'''
            <a href="{reverse('admin:orders_shippingdocument_change', args=[obj.id])}" 
               style="background: #3498db; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
               ✏️
            </a>
            <a href="{obj.file.url}" target="_blank" 
               style="background: #27ae60; color: white; padding: 4px 8px; border-radius: 5px; text-decoration: none; font-size: 11px;">
               📎
            </a>
        ''')
    action_buttons.short_description = 'عملیات'