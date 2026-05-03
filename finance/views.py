from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum
from orders.models import Order, Customer, Driver, Vehicle
from finance.models import Invoice
from common.decorators import role_required
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone

@login_required
@role_required('can_view_reports')
def finance_dashboard(request):
    """داشبورد مالی با آمار"""
    
    total_orders = Order.objects.count()
    total_customers = Customer.objects.count()
    total_drivers = Driver.objects.count()
    total_vehicles = Vehicle.objects.count()
    
    total_invoices = Invoice.objects.aggregate(total=Sum('total'))['total'] or 0
    paid_invoices = Invoice.objects.filter(status='paid').aggregate(total=Sum('total'))['total'] or 0
    pending_invoices = total_invoices - paid_invoices
    
    # سفارشات این ماه
    now = timezone.now()
    monthly_orders = Order.objects.filter(
        created_at__year=now.year,
        created_at__month=now.month
    ).count()
    
    context = {
        'total_orders': total_orders,
        'total_customers': total_customers,
        'total_drivers': total_drivers,
        'total_vehicles': total_vehicles,
        'total_invoices': total_invoices,
        'paid_invoices': paid_invoices,
        'pending_invoices': pending_invoices,
        'monthly_orders': monthly_orders,
        'currency_symbol': '؋',
        'currency_name': 'افغانی',
    }
    return render(request, 'finance/dashboard_report.html', context)


@login_required
@role_required('can_view_reports')
def export_orders_excel(request):
    """خروجی اکسل سفارشات"""
    orders = Order.objects.all().select_related('customer', 'driver')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="orders_export.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سفارشات"
    
    headers = ['شماره سفارش', 'مشتری', 'مبدا', 'مقصد', 'نوع بار', 'وزن', 'مبلغ (؋)', 'وضعیت', 'تاریخ ثبت']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=order.customer.name)
        ws.cell(row=row, column=3, value=order.origin)
        ws.cell(row=row, column=4, value=order.destination)
        ws.cell(row=row, column=5, value=order.cargo_type)
        ws.cell(row=row, column=6, value=float(order.weight))
        ws.cell(row=row, column=7, value=float(order.price))
        ws.cell(row=row, column=8, value=order.get_status_display())
        ws.cell(row=row, column=9, value=order.created_at.strftime('%Y/%m/%d'))
    
    # تنظیم عرض ستون‌ها
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(response)
    return response


@login_required
@role_required('can_view_reports')
def export_invoices_excel(request):
    """خروجی اکسل فاکتورها"""
    invoices = Invoice.objects.all().select_related('customer', 'order')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="invoices_export.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "فاکتورها"
    
    headers = ['شماره فاکتور', 'مشتری', 'سفارش', 'مبلغ کل', 'تخفیف', 'قابل پرداخت', 'وضعیت', 'سررسید']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, invoice in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=invoice.invoice_number)
        ws.cell(row=row, column=2, value=invoice.customer.name)
        ws.cell(row=row, column=3, value=invoice.order.order_number if invoice.order else '-')
        ws.cell(row=row, column=4, value=float(invoice.subtotal))
        ws.cell(row=row, column=5, value=float(invoice.discount))
        ws.cell(row=row, column=6, value=float(invoice.total))
        ws.cell(row=row, column=7, value=invoice.get_status_display())
        ws.cell(row=row, column=8, value=invoice.due_date.strftime('%Y/%m/%d') if invoice.due_date else '-')
    
    # تنظیم عرض ستون‌ها
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(response)
    return response